import re
import shutil
from pathlib import Path
from openpyxl import load_workbook


# =====================================================
# Parse markdown tables from testcases.txt
# =====================================================

def parse_testcases(txt_file):

    with open(txt_file, "r", encoding="utf-8") as f:
        content = f.read()

    testcases = {}
    current_headers = None

    for line in content.splitlines():

        line = line.strip()

        if not line.startswith("|"):
            continue

        # Header row
        if (
            "TC-ID" in line
            and "Test Scenario" in line
            and "Expected Result" in line
        ):
            current_headers = [
                h.strip()
                for h in line.strip("|").split("|")
            ]
            continue

        # Separator row
        if re.match(r"^\|\s*-+", line):
            continue

        if not current_headers:
            continue

        values = [
            v.strip()
            for v in line.strip("|").split("|")
        ]

        if len(values) != len(current_headers):
            continue

        row = dict(zip(current_headers, values))

        tc_id = row.get("TC-ID", "").strip()

        if not tc_id:
            continue

        testcases[tc_id] = {
            "Scenario": row.get("Test Scenario", ""),
            "Steps": row.get("Test Steps", "").replace("<br>", "\n"),
            "Expected Result": row.get("Expected Result", ""),
            "Priority": row.get("Priority", "")
        }

    return testcases


# =====================================================
# Populate Excel
# =====================================================

def populate_excel(excel_file, testcase_data):

    wb = load_workbook(excel_file)

    total_updated = 0

    for ws in wb.worksheets:

        print(f"\nProcessing Sheet: {ws.title}")

        sheet_updated = 0

        for row in range(1, ws.max_row + 1):

            tc_id = ws.cell(row, 1).value

            if tc_id is None:
                continue

            tc_id = str(tc_id).strip()

            if tc_id not in testcase_data:
                continue

            data = testcase_data[tc_id]

            # =================================================
            # Sheet with 8+ columns
            # A = TC ID
            # B = Scenario
            # D = Steps
            # F = Expected Result
            # H = Priority
            # =================================================

            if ws.max_column >= 8:

                ws.cell(row, 2).value = data["Scenario"]
                ws.cell(row, 4).value = data["Steps"]
                ws.cell(row, 6).value = data["Expected Result"]
                ws.cell(row, 8).value = data["Priority"]

            # =================================================
            # Smaller sheet
            # A = TC ID
            # B = Scenario
            # C = Priority
            # =================================================

            elif ws.max_column >= 3:

                ws.cell(row, 2).value = data["Scenario"]
                ws.cell(row, 3).value = data["Priority"]

            sheet_updated += 1
            total_updated += 1

        print(f"Updated {sheet_updated} test cases")

    wb.save(excel_file)

    return total_updated


# =====================================================
# Main
# =====================================================

def main():

    base_dir = Path(__file__).resolve().parent

    txt_file = base_dir / "testcases.txt"
    template_file = base_dir / "template.xlsx"

    output_dir = base_dir / "output"
    output_dir.mkdir(exist_ok=True)

    output_file = output_dir / "output.xlsx"

    print("=" * 60)
    print("Reading Input Files")
    print("=" * 60)

    if not txt_file.exists():
        raise FileNotFoundError(
            f"testcases.txt not found:\n{txt_file}"
        )

    if not template_file.exists():
        raise FileNotFoundError(
            f"template.xlsx not found:\n{template_file}"
        )

    # Create output workbook from template
    if not output_file.exists():

        shutil.copy(
            template_file,
            output_file
        )

        print(f"Created output file:")
        print(output_file)

    testcase_data = parse_testcases(txt_file)

    print(f"\nParsed {len(testcase_data)} test cases")

    print("\n" + "=" * 60)
    print("Updating Excel")
    print("=" * 60)

    updated = populate_excel(
        output_file,
        testcase_data
    )

    print("\n" + "=" * 60)
    print("Completed")
    print("=" * 60)
    print(f"Updated Test Cases : {updated}")
    print(f"Saved File         : {output_file}")


if __name__ == "__main__":
    main()


    